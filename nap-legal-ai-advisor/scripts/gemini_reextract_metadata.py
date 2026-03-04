#!/usr/bin/env python3
"""
Gemini-powered metadata re-extraction and validation.

For each decision in cleaned_court_texts.json, sends key sections to Gemini
for structured metadata extraction, then compares against current metadata
and produces:
  - Data/validation/gemini_reextraction_report.xlsx  (comparison + summary)
  - Data/validation/gemini_corrections.json          (diffs only)

Usage:
    export GEMINI_API_KEY=AIzaSyAm3RYdgNW-OEQXDNSYN7t4GGuC06_8MzQ
    cd nap-legal-ai-advisor
    python scripts/gemini_reextract_metadata.py --data-root ../Data
"""

import argparse
import json
import os
import re
import ssl
import sys
import time
import urllib.error
import urllib.request
from datetime import datetime, timezone

import google.generativeai as genai
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

EXTRACTION_SYSTEM_PROMPT = """Du är en juridisk dataextraktionsexpert specialiserad på svenska miljödomstolsbeslut.

Du får texten från ett domstolsbeslut. Din uppgift är att extrahera strukturerad metadata.

Svara ENBART med ett JSON-objekt med följande fält (använd null om information saknas):

{
  "is_nap_hydropower": true/false,
  "is_nap_reason": "kort förklaring om det INTE är NAP/vattenkraft",
  "power_plant_name": "kraftverkets namn (utan 'moderna miljövillkor avseende/vid/för')",
  "operator_name": "verksamhetsutövarens namn",
  "watercourse": "vattendragets namn (å, älv, sjö)",
  "municipality": "kommunnamn",
  "application_outcome": "en av: granted, granted_modified, conditions_changed, denied, appeal_denied, remanded, overturned, dismissed, withdrawn",
  "application_outcome_sv": "svensk beskrivning av utfallet",
  "outcome_reasoning": "1 mening som förklarar utfallet",
  "measures_ordered": ["lista", "av", "åtgärder", "som", "domstolen", "beslutat"],
  "operator_cost_sek": null eller ett tal (BARA kostnader som åläggs verksamhetsutövaren — EJ miljöskadeberäkningar, EJ samhällskostnader, EJ kompensationskrav från tredje part),
  "cost_description": "kort beskrivning av vad kostnaden avser",
  "environmental_measures": ["fiskväg", "minimitappning", etc — specifika miljöåtgärder]
}

VIKTIGA REGLER:
- is_nap_hydropower: TRUE om beslutet handlar om vattenkraft, dammar, vattenreglering, eller moderna miljövillkor för vattenverksamhet. FALSE om det handlar om bergtäkt, solceller, detaljplan, utsläppsrätter, eller annat som inte rör vattenkraft.
- power_plant_name: Extrahera BARA kraftverkets/dammens/kvarns namn. INTE "moderna miljövillkor avseende X" — bara "X".
- operator_cost_sek: BARA den faktiska kostnaden som verksamhetsutövaren tvingas betala. Inte miljöskadeberäkningar som nämns i domskäl, inte samhällskostnader, inte tredjepartskrav. Om du är osäker, sätt null.
- application_outcome:
  "granted" = tillstånd beviljat som ansökt
  "granted_modified" = tillstånd beviljat med ändringar
  "conditions_changed" = villkor ändras/uppdateras
  "denied" = ansökan avslås
  "appeal_denied" = överklagande avslås (överklagande avvisas)
  "remanded" = återförvisas till lägre instans
  "overturned" = lägre instans beslut upphävs
  "dismissed" = avvisas på formella grunder
  "withdrawn" = återkallat av sökanden
- measures_ordered: lista bara åtgärder som DOMSTOLEN BESLUTAT, inte vad parterna yrkat
- Svara ENBART med JSON, ingen annan text
"""

GEMINI_MODELS = [
    "gemini-2.5-flash",
    "gemini-2.5-flash-lite",
    "gemini-2.0-flash",
    "gemini-2.0-flash-lite",
]
GEMINI_MODEL = GEMINI_MODELS[0]  # default / reported in output

GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
HEADER_FONT = Font(bold=True)
WRAP_ALIGNMENT = Alignment(wrap_text=True, vertical="top")

# ---------------------------------------------------------------------------
# Path resolution (same logic as other scripts in this directory)
# ---------------------------------------------------------------------------


def resolve_data_root(cli_arg=None):
    """Resolve the Data directory root."""
    if cli_arg:
        p = os.path.abspath(cli_arg)
        if os.path.isdir(p):
            return p
        print(f"ERROR: --data-root '{cli_arg}' is not a directory")
        sys.exit(1)

    script_dir = os.path.dirname(os.path.abspath(__file__))
    marker = os.path.join("Data", "processed", "cleaned_court_texts.json")

    candidates = [
        os.getcwd(),
        os.path.join(script_dir, ".."),
        os.path.join(script_dir, "..", ".."),
        os.path.join(script_dir, "..", "..", ".."),
    ]
    for base in candidates:
        check = os.path.join(base, marker)
        if os.path.isfile(check):
            return os.path.abspath(os.path.join(base, "Data"))

    print("ERROR: Could not locate Data/processed/cleaned_court_texts.json")
    print("Use --data-root to specify the Data directory explicitly.")
    sys.exit(1)


# ---------------------------------------------------------------------------
# Gemini extraction
# ---------------------------------------------------------------------------


def build_user_prompt(decision):
    """Build the user-message prompt from a decision object."""
    sections = decision.get("sections", {})
    domslut = sections.get("domslut", "")
    bakgrund = sections.get("bakgrund", "") or sections.get("saken", "")
    domskal = sections.get("domskäl", "")

    # Fallback to full_text if no sections available
    if not domslut and not domskal:
        domslut = decision.get("text_full", "")[:6000]

    meta = decision.get("metadata", {})
    case_number = meta.get("case_number", decision.get("id", ""))
    court = meta.get("court", "")
    date = meta.get("date", "")

    return f"""Extrahera metadata från följande domstolsbeslut:

Målnummer: {case_number}
Domstol: {court}
Datum: {date}

--- DOMSLUT ---
{domslut[:4000]}

--- BAKGRUND/SAKEN ---
{bakgrund[:2000]}

--- DOMSKÄL (utdrag) ---
{domskal[:2000]}
"""


def _gemini_rest_call(api_key, user_text, model_name=None):
    """Call Gemini REST API directly, bypassing gRPC (avoids corporate proxy SSL issues)."""
    model_name = model_name or GEMINI_MODEL
    url = (
        f"https://generativelanguage.googleapis.com/v1beta/models/"
        f"{model_name}:generateContent?key={api_key}"
    )
    payload = {
        "contents": [{"role": "user", "parts": [{"text": user_text}]}],
        "generationConfig": {
            "responseMimeType": "application/json",
            "temperature": 0.1,
        },
    }
    data = json.dumps(payload).encode("utf-8")

    # Build an SSL context that skips cert verification (corporate proxy workaround)
    ctx = ssl.create_default_context()
    ctx.check_hostname = False
    ctx.verify_mode = ssl.CERT_NONE

    req = urllib.request.Request(
        url, data=data, headers={"Content-Type": "application/json"}, method="POST"
    )
    with urllib.request.urlopen(req, context=ctx, timeout=120) as resp:
        body = json.loads(resp.read().decode("utf-8"))

    # Extract text from Gemini response
    return body["candidates"][0]["content"]["parts"][0]["text"]


def extract_metadata_gemini(api_key, decision, max_retries=5):
    """Call Gemini to extract metadata for one decision. Returns parsed dict or None.

    Cycles through GEMINI_MODELS on 429 rate-limit errors before retrying.
    """
    prompt = build_user_prompt(decision)
    full_prompt = EXTRACTION_SYSTEM_PROMPT + "\n\n" + prompt

    for attempt in range(1 + max_retries):
        # Pick a model: cycle through the list on successive retries
        model_name = GEMINI_MODELS[attempt % len(GEMINI_MODELS)]

        try:
            text = _gemini_rest_call(api_key, full_prompt, model_name).strip()

            # Strip markdown code fences if present
            if text.startswith("```"):
                text = text.split("\n", 1)[1]
            if text.endswith("```"):
                text = text.rsplit("```", 1)[0]
            text = text.strip()

            return json.loads(text)

        except json.JSONDecodeError as e:
            if attempt < max_retries:
                time.sleep(5)
                continue
            print(f"(JSON parse error: {e}) ", end="", flush=True)
            return None
        except urllib.error.HTTPError as e:
            if e.code == 429 and attempt < max_retries:
                next_model = GEMINI_MODELS[(attempt + 1) % len(GEMINI_MODELS)]
                wait = min(5 * (2 ** (attempt // len(GEMINI_MODELS))), 30)
                print(f"(429@{model_name}, try {next_model} in {wait}s) ", end="", flush=True)
                time.sleep(wait)
                continue
            if attempt < max_retries:
                time.sleep(5)
                continue
            print(f"(API error: {e}) ", end="", flush=True)
            return None
        except Exception as e:
            if attempt < max_retries:
                time.sleep(5)
                continue
            print(f"(API error: {e}) ", end="", flush=True)
            return None

    return None


# ---------------------------------------------------------------------------
# Comparison helpers
# ---------------------------------------------------------------------------


def _normalize(s):
    """Lowercase, strip whitespace and common suffixes for fuzzy comparison."""
    if s is None:
        return ""
    s = str(s).strip().lower()
    # Remove common suffixes that don't affect semantic match
    for suffix in ["kraftverk", "kvarn", "kraftstation", "vattenkraftverk"]:
        if s.endswith(suffix):
            s = s[: -len(suffix)].strip()
    # Remove trailing punctuation and whitespace
    s = s.rstrip(" -,.")
    return s


def names_match(a, b):
    """Check if two names are essentially the same."""
    na, nb = _normalize(a), _normalize(b)
    if not na and not nb:
        return True
    if not na or not nb:
        return False
    return na == nb or na in nb or nb in na


def costs_match(current, gemini):
    """Both null → True, otherwise within 20%."""
    if current is None and gemini is None:
        return True
    if current is None or gemini is None:
        return False
    try:
        c, g = float(current), float(gemini)
    except (ValueError, TypeError):
        return False
    if c == 0 and g == 0:
        return True
    if c == 0 or g == 0:
        return False
    ratio = abs(c - g) / max(abs(c), abs(g))
    return ratio <= 0.20


def outcomes_match(current, gemini):
    """Simple case-insensitive equality for outcome strings."""
    if not current and not gemini:
        return True
    if not current or not gemini:
        return False
    return str(current).strip().lower() == str(gemini).strip().lower()


def measures_to_str(measures):
    """Convert measures list (could be list of dicts or list of strings) to comma-separated string."""
    if not measures:
        return ""
    parts = []
    for m in measures:
        if isinstance(m, dict):
            t = m.get("type", "")
            d = m.get("details", "")
            parts.append(f"{t}: {d}" if d else t)
        else:
            parts.append(str(m))
    return ", ".join(parts)


# ---------------------------------------------------------------------------
# Excel report generation
# ---------------------------------------------------------------------------

COMPARISON_HEADERS = [
    "case_number",
    "date",
    "court",
    "GEMINI_is_nap",
    "GEMINI_nap_reason",
    "CURRENT_plant_name",
    "GEMINI_plant_name",
    "PLANT_MATCH",
    "CURRENT_operator",
    "GEMINI_operator",
    "OPERATOR_MATCH",
    "CURRENT_watercourse",
    "GEMINI_watercourse",
    "WATERCOURSE_MATCH",
    "CURRENT_outcome",
    "GEMINI_outcome",
    "OUTCOME_MATCH",
    "GEMINI_outcome_reasoning",
    "CURRENT_cost",
    "GEMINI_cost",
    "COST_MATCH",
    "GEMINI_cost_description",
    "CURRENT_measures",
    "GEMINI_measures",
    "GEMINI_environmental_measures",
]

MATCH_COLUMNS = {"PLANT_MATCH", "OPERATOR_MATCH", "WATERCOURSE_MATCH", "OUTCOME_MATCH", "COST_MATCH"}


def build_comparison_row(decision, gemini_data):
    """Build a single comparison row dict."""
    meta = decision.get("metadata", {})

    current_plant = meta.get("power_plant_name")
    gemini_plant = gemini_data.get("power_plant_name")

    current_operator = meta.get("operator_name")
    gemini_operator = gemini_data.get("operator_name")

    current_watercourse = meta.get("watercourse")
    gemini_watercourse = gemini_data.get("watercourse")

    current_outcome = meta.get("application_outcome")
    gemini_outcome = gemini_data.get("application_outcome")

    current_cost = meta.get("total_cost_sek")
    gemini_cost = gemini_data.get("operator_cost_sek")

    current_measures = meta.get("measures_ordered") or []
    gemini_measures = gemini_data.get("measures_ordered") or []
    gemini_env_measures = gemini_data.get("environmental_measures") or []

    return {
        "case_number": meta.get("case_number", decision.get("id", "")),
        "date": meta.get("date", ""),
        "court": meta.get("court", ""),
        "GEMINI_is_nap": gemini_data.get("is_nap_hydropower"),
        "GEMINI_nap_reason": gemini_data.get("is_nap_reason"),
        "CURRENT_plant_name": current_plant,
        "GEMINI_plant_name": gemini_plant,
        "PLANT_MATCH": names_match(current_plant, gemini_plant),
        "CURRENT_operator": current_operator,
        "GEMINI_operator": gemini_operator,
        "OPERATOR_MATCH": names_match(current_operator, gemini_operator),
        "CURRENT_watercourse": current_watercourse,
        "GEMINI_watercourse": gemini_watercourse,
        "WATERCOURSE_MATCH": names_match(current_watercourse, gemini_watercourse),
        "CURRENT_outcome": current_outcome,
        "GEMINI_outcome": gemini_outcome,
        "OUTCOME_MATCH": outcomes_match(current_outcome, gemini_outcome),
        "GEMINI_outcome_reasoning": gemini_data.get("outcome_reasoning"),
        "CURRENT_cost": current_cost,
        "GEMINI_cost": gemini_cost,
        "COST_MATCH": costs_match(current_cost, gemini_cost),
        "GEMINI_cost_description": gemini_data.get("cost_description"),
        "CURRENT_measures": measures_to_str(current_measures),
        "GEMINI_measures": measures_to_str(gemini_measures),
        "GEMINI_environmental_measures": measures_to_str(gemini_env_measures),
    }


def write_comparison_sheet(ws, rows):
    """Write Sheet 1: Jämförelse."""
    # Headers
    for col_idx, header in enumerate(COMPARISON_HEADERS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.alignment = WRAP_ALIGNMENT

    # Data rows
    for row_idx, row_data in enumerate(rows, 2):
        for col_idx, header in enumerate(COMPARISON_HEADERS, 1):
            value = row_data.get(header)
            # Convert booleans for display
            if isinstance(value, bool):
                value = "TRUE" if value else "FALSE"
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = WRAP_ALIGNMENT

            # Apply conditional formatting
            if header in MATCH_COLUMNS:
                cell.fill = GREEN_FILL if row_data.get(header) else RED_FILL
            if header == "GEMINI_is_nap":
                if row_data.get("GEMINI_is_nap") is False:
                    cell.fill = RED_FILL

    # Freeze header row
    ws.freeze_panes = "A2"

    # Auto-filter
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COMPARISON_HEADERS))}1"

    # Auto-width (cap at 40)
    for col_idx in range(1, len(COMPARISON_HEADERS) + 1):
        max_len = len(str(ws.cell(row=1, column=col_idx).value or ""))
        for row_idx in range(2, len(rows) + 2):
            val = str(ws.cell(row=row_idx, column=col_idx).value or "")
            max_len = max(max_len, min(len(val), 40))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 42)


def write_summary_sheet(ws, rows, failed_count):
    """Write Sheet 2: Sammanfattning."""
    total = len(rows)

    # Non-NAP decisions
    non_nap = [r for r in rows if r.get("GEMINI_is_nap") is False]
    non_nap_cases = [r["case_number"] for r in non_nap]

    # Match rates
    def match_rate(field):
        if total == 0:
            return 0
        return sum(1 for r in rows if r.get(field)) / total * 100

    plant_rate = match_rate("PLANT_MATCH")
    operator_rate = match_rate("OPERATOR_MATCH")
    watercourse_rate = match_rate("WATERCOURSE_MATCH")
    outcome_rate = match_rate("OUTCOME_MATCH")
    cost_rate = match_rate("COST_MATCH")

    # Disagreements
    outcome_disagree = [
        f"{r['case_number']}: {r.get('CURRENT_outcome')} -> {r.get('GEMINI_outcome')}"
        for r in rows
        if not r.get("OUTCOME_MATCH")
    ]
    cost_disagree = [
        f"{r['case_number']}: {r.get('CURRENT_cost')} -> {r.get('GEMINI_cost')}"
        for r in rows
        if not r.get("COST_MATCH")
    ]

    lines = [
        f"Total decisions processed: {total}",
        f"Gemini API failures: {failed_count}",
        "",
        f"Non-NAP flagged by Gemini: {len(non_nap)} ({', '.join(non_nap_cases) if non_nap_cases else 'none'})",
        "",
        "Field match rates:",
        f"  Plant name:  {plant_rate:.1f}% match",
        f"  Operator:    {operator_rate:.1f}% match",
        f"  Watercourse: {watercourse_rate:.1f}% match",
        f"  Outcome:     {outcome_rate:.1f}% match",
        f"  Cost:        {cost_rate:.1f}% match",
        "",
        f"Decisions with outcome disagreement ({len(outcome_disagree)}):",
    ]
    for d in outcome_disagree:
        lines.append(f"  {d}")

    lines.append("")
    lines.append(f"Decisions with cost disagreement ({len(cost_disagree)}):")
    for d in cost_disagree:
        lines.append(f"  {d}")

    for row_idx, line in enumerate(lines, 1):
        cell = ws.cell(row=row_idx, column=1, value=line)
        if row_idx == 1 or line.startswith("Field match") or line.startswith("Decisions with"):
            cell.font = HEADER_FONT

    ws.column_dimensions["A"].width = 80


def write_raw_data_sheet(ws, raw_data):
    """Write Sheet 3: Gemini_rådata."""
    ws.cell(row=1, column=1, value="case_number").font = HEADER_FONT
    ws.cell(row=1, column=2, value="full_json_response").font = HEADER_FONT

    for row_idx, (case_number, json_response) in enumerate(raw_data, 2):
        ws.cell(row=row_idx, column=1, value=case_number)
        ws.cell(row=row_idx, column=2, value=json_response)

    ws.freeze_panes = "A2"
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 80


# ---------------------------------------------------------------------------
# Corrections JSON generation
# ---------------------------------------------------------------------------


def build_corrections(decisions_map, results):
    """Build corrections list — only decisions where Gemini differs from current metadata."""
    corrections = []

    for decision_id, gemini_data in results.items():
        decision = decisions_map[decision_id]
        meta = decision.get("metadata", {})
        fields_changed = {}

        # Field mapping: (current_key, gemini_key)
        field_pairs = [
            ("power_plant_name", "power_plant_name"),
            ("operator_name", "operator_name"),
            ("watercourse", "watercourse"),
            ("application_outcome", "application_outcome"),
            ("application_outcome_sv", "application_outcome_sv"),
        ]

        for current_key, gemini_key in field_pairs:
            current_val = meta.get(current_key)
            gemini_val = gemini_data.get(gemini_key)

            # Normalize for comparison
            c_norm = (str(current_val).strip().lower() if current_val else "")
            g_norm = (str(gemini_val).strip().lower() if gemini_val else "")

            if c_norm != g_norm and (current_val or gemini_val):
                fields_changed[current_key] = {
                    "old": current_val,
                    "new": gemini_val,
                    "source": "gemini",
                }

        # Cost comparison
        current_cost = meta.get("total_cost_sek")
        gemini_cost = gemini_data.get("operator_cost_sek")
        if not costs_match(current_cost, gemini_cost):
            fields_changed["total_cost_sek"] = {
                "old": current_cost,
                "new": gemini_cost,
                "source": "gemini",
            }

        if fields_changed:
            corrections.append(
                {
                    "id": decision.get("id", ""),
                    "case_number": meta.get("case_number", decision.get("id", "")),
                    "fields_changed": fields_changed,
                    "is_nap": gemini_data.get("is_nap_hydropower", True),
                    "gemini_full_response": gemini_data,
                }
            )

    return corrections


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------


def main():
    parser = argparse.ArgumentParser(
        description="Gemini-powered metadata re-extraction and comparison"
    )
    parser.add_argument(
        "--data-root",
        default=None,
        help="Path to the Data directory (default: auto-detect)",
    )
    parser.add_argument(
        "--max-consecutive-failures",
        type=int,
        default=3,
        help="Stop API calls after N consecutive failures (quota exhausted). "
             "Cached results are still used for the report. (default: 3)",
    )
    args = parser.parse_args()

    # --- Resolve paths ---
    data_root = resolve_data_root(args.data_root)
    input_path = os.path.join(data_root, "processed", "cleaned_court_texts.json")
    validation_dir = os.path.join(data_root, "validation")
    os.makedirs(validation_dir, exist_ok=True)

    report_path = os.path.join(validation_dir, "gemini_reextraction_report.xlsx")
    corrections_path = os.path.join(validation_dir, "gemini_corrections.json")

    # --- Check API key ---
    api_key = os.environ.get("GEMINI_API_KEY")
    if not api_key:
        print("ERROR: GEMINI_API_KEY environment variable is not set.")
        print('  export GEMINI_API_KEY="your-key-here"')
        sys.exit(1)

    # REST API is used directly (bypasses gRPC SSL issues with corporate proxies)

    # --- Load decisions ---
    print(f"Loading decisions from {input_path} ...")
    with open(input_path, "r", encoding="utf-8") as f:
        data = json.load(f)

    decisions = data.get("decisions", [])
    total = len(decisions)
    print(f"Found {total} decisions.\n")

    if total == 0:
        print("No decisions to process.")
        sys.exit(0)

    # Build lookup map
    decisions_map = {}
    for d in decisions:
        did = d.get("id", "")
        decisions_map[did] = d

    # --- Load cache for resume support ---
    cache_path = os.path.join(validation_dir, "_gemini_cache.json")
    cache = {}
    if os.path.isfile(cache_path):
        try:
            with open(cache_path, "r", encoding="utf-8") as f:
                cache = json.load(f)
            print(f"Resuming: {len(cache)} decisions already cached.\n")
        except Exception:
            cache = {}

    # --- Process each decision ---
    results = {}  # decision_id -> gemini_data
    raw_responses = []  # (case_number, json_string)
    comparison_rows = []
    succeeded = 0
    failed = 0
    cached = 0
    consecutive_fail = 0

    for idx, decision in enumerate(decisions, 1):
        did = decision.get("id", "")
        meta = decision.get("metadata", {})
        case_number = meta.get("case_number", did)
        was_cached = did in cache

        # Use cached result if available
        if was_cached:
            gemini_data = cache[did]
            print(f"[{idx}/{total}] {case_number}... CACHED")
            cached += 1
        else:
            # Stop making API calls if quota is clearly exhausted
            if consecutive_fail >= args.max_consecutive_failures:
                print(f"[{idx}/{total}] {case_number}... SKIPPED (quota exhausted)")
                failed += 1
                continue

            # If we've hit 2+ consecutive failures, do a long cooldown
            if consecutive_fail >= 2:
                cooldown = 60
                print(f"  ** Quota cooldown: waiting {cooldown}s before next attempt...")
                time.sleep(cooldown)

            print(f"[{idx}/{total}] {case_number}...", end=" ", flush=True)
            gemini_data = extract_metadata_gemini(api_key, decision)

            if gemini_data is None:
                print("FAIL (extraction failed)")
                failed += 1
                consecutive_fail += 1
                continue

            print("OK")
            consecutive_fail = 0

            # Save to cache immediately
            cache[did] = gemini_data
            with open(cache_path, "w", encoding="utf-8") as f:
                json.dump(cache, f, ensure_ascii=False)

        succeeded += 1
        results[did] = gemini_data

        # Store raw response
        raw_responses.append((case_number, json.dumps(gemini_data, ensure_ascii=False, indent=2)))

        # Build comparison row
        row = build_comparison_row(decision, gemini_data)
        comparison_rows.append(row)

        # Rate limiting — Gemini free tier allows ~10 RPM
        if not was_cached and idx < total:
            time.sleep(6)

    # --- Summary ---
    print(f"\n{'='*50}")
    print(f"Extraction complete: {succeeded} succeeded ({cached} cached), {failed} failed")
    print(f"{'='*50}\n")

    # --- Generate Excel report ---
    print("Generating Excel report ...")
    wb = Workbook()

    # Sheet 1: Jämförelse
    ws1 = wb.active
    ws1.title = "Jämförelse"
    write_comparison_sheet(ws1, comparison_rows)

    # Sheet 2: Sammanfattning
    ws2 = wb.create_sheet("Sammanfattning")
    write_summary_sheet(ws2, comparison_rows, failed)

    # Sheet 3: Gemini_rådata
    ws3 = wb.create_sheet("Gemini_rådata")
    write_raw_data_sheet(ws3, raw_responses)

    wb.save(report_path)
    print(f"  -> {report_path}")

    # --- Generate corrections JSON ---
    print("Generating corrections JSON ...")
    corrections = build_corrections(decisions_map, results)
    corrections_output = {
        "generated": datetime.now(timezone.utc).isoformat(),
        "model": ", ".join(GEMINI_MODELS),
        "total_decisions_processed": succeeded,
        "total_corrections": len(corrections),
        "corrections": corrections,
    }

    with open(corrections_path, "w", encoding="utf-8") as f:
        json.dump(corrections_output, f, ensure_ascii=False, indent=2)
    print(f"  -> {corrections_path}")

    print(f"\nDone. {len(corrections)} decisions have differing metadata.")


if __name__ == "__main__":
    main()
