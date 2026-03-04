#!/usr/bin/env python3
"""
Generate data validation report for 50 NAP court decisions.

Outputs: Data/validation/data_validation_report.xlsx
  - Sheet 1: Datavalidering (one row per decision)
  - Sheet 2: Sammanfattning (summary statistics)
  - Sheet 3: NAP_landskapet (NAP landscape context)
"""

import argparse
import csv
import json
import os
import sys
from datetime import datetime

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: openpyxl is required. Install with: pip install openpyxl")
    sys.exit(1)


# ---------------------------------------------------------------------------
# Path resolution
# ---------------------------------------------------------------------------

def resolve_data_root(cli_arg=None):
    """Resolve the Data directory root."""
    if cli_arg:
        p = os.path.abspath(cli_arg)
        if os.path.isdir(p):
            return p
        print(f"ERROR: --data-root '{cli_arg}' is not a directory")
        sys.exit(1)

    script_dir = os.path.dirname(os.path.abspath(__file__))
    marker = os.path.join("Data", "processed", "cleaned_court_texts.json")

    # Search upward from cwd, then from script dir
    candidates = [
        os.getcwd(),
        os.path.join(script_dir, ".."),
        os.path.join(script_dir, "..", ".."),
        os.path.join(script_dir, "..", "..", ".."),
    ]
    for base in candidates:
        check = os.path.join(base, marker)
        if os.path.isfile(check):
            return os.path.abspath(os.path.join(base, "Data"))

    print("ERROR: Could not locate Data/processed/cleaned_court_texts.json")
    print("Use --data-root to specify the Data directory explicitly.")
    sys.exit(1)


# ---------------------------------------------------------------------------
# Data loading helpers
# ---------------------------------------------------------------------------

def load_json(path):
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)


def load_strommen_csv(path):
    """Load a Strommen CSV file (semicolon-delimited, UTF-8 with BOM)."""
    rows = []
    with open(path, "r", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f, delimiter=";")
        for row in reader:
            rows.append(row)
    return rows


def normalize_case_number(cn):
    """Normalize case number: strip, remove spaces/dashes, uppercase."""
    if not cn:
        return ""
    return cn.strip().replace(" ", "").replace("-", "").upper()


def build_label_lookup(labeled_data):
    """Build lookup {id: {label, scoring_details}} from labeled_dataset_binary."""
    lookup = {}
    for split_name in ("train", "val", "test"):
        items = labeled_data.get("splits", {}).get(split_name, [])
        for item in items:
            lookup[item["id"]] = {
                "label": item["label"],
                "scoring_details": item.get("scoring_details", {}),
            }
    return lookup


def build_strommen_lookup(strommen_rows):
    """Build lookup by normalized case number from Strommen data.

    If multiple plants share the same case number (provningsgrupp), store the
    FIRST match and note the plant count.
    """
    lookup = {}  # normalized_cn -> {row, plant_count}
    cn_counts = {}

    for row in strommen_rows:
        for col in ("Målnummer", "Målnummer MMÖD"):
            cn_raw = (row.get(col) or "").strip()
            if not cn_raw:
                continue
            ncn = normalize_case_number(cn_raw)
            if not ncn:
                continue
            cn_counts[ncn] = cn_counts.get(ncn, 0) + 1
            if ncn not in lookup:
                lookup[ncn] = {"row": row, "plant_count": 1}

    # Update plant counts
    for ncn in lookup:
        if ncn in cn_counts:
            lookup[ncn]["plant_count"] = cn_counts[ncn]

    return lookup


# ---------------------------------------------------------------------------
# Outcome alignment check
# ---------------------------------------------------------------------------

def check_outcome_alignment(our_outcome, str_beslut):
    """Check if our application_outcome aligns with Strommen beslut.

    Returns True if they ALIGN (no flag needed), False if MISALIGNED.
    """
    if not str_beslut or str_beslut.strip() == "Ej Beslut":
        return True  # No decision to compare against

    beslut = str_beslut.strip()
    outcome = (our_outcome or "").strip()

    if beslut == "Avvisning":
        return outcome in ("denied", "appeal_denied", "unclear")
    elif beslut == "Moderna miljövillkor":
        return outcome in ("granted", "granted_modified", "conditions_changed")
    elif beslut == "Avslag":
        return outcome in ("denied",)
    elif beslut.startswith("Återkallande"):
        # Special category we may not have — don't flag
        return True
    elif beslut == "Avskrivet":
        return True  # Don't flag

    # If our outcome is unclear and STR has any specific beslut, flag it
    if outcome == "unclear":
        return False

    return True


# ---------------------------------------------------------------------------
# Validation flags
# ---------------------------------------------------------------------------

def compute_flags(row_data, str_match, label_info):
    """Compute all FLAG_ columns and return as dict."""
    flags = {}

    total_cost = row_data.get("total_cost_sek")
    flags["FLAG_cost_over_5M"] = (
        total_cost is not None and total_cost > 5_000_000
    )

    flags["FLAG_outcome_unclear"] = (
        row_data.get("application_outcome") == "unclear"
    )

    measures_ordered_str = row_data.get("measures_ordered_str", "")
    domslut_measures_str = row_data.get("domslut_measures_str", "")
    extracted_measures_str = row_data.get("extracted_measures_str", "")
    flags["FLAG_no_measures"] = (
        not measures_ordered_str and not domslut_measures_str
        and not extracted_measures_str
    )

    flags["FLAG_no_cost"] = total_cost is None or total_cost == 0

    plant = row_data.get("power_plant_name")
    flags["FLAG_no_plant_name"] = (
        plant is None or str(plant).strip() == "" or str(plant) == "None"
    )

    operator = row_data.get("operator_name")
    flags["FLAG_no_operator"] = (
        operator is None or str(operator).strip() == "" or str(operator) == "None"
    )

    flags["FLAG_no_processing_time"] = (
        row_data.get("processing_time_days") is None
    )

    flags["FLAG_unlabeled"] = row_data.get("risk_label", "") == ""

    plant_str = str(plant or "").lower()
    flags["FLAG_plant_name_prefix"] = "moderna miljövillkor" in plant_str

    # Outcome mismatch with Strommen
    if str_match:
        str_beslut = str_match["row"].get("Beslut", "")
        aligned = check_outcome_alignment(
            row_data.get("application_outcome"), str_beslut
        )
        flags["FLAG_outcome_mismatch"] = not aligned
    else:
        flags["FLAG_outcome_mismatch"] = False

    return flags


# ---------------------------------------------------------------------------
# Report generation
# ---------------------------------------------------------------------------

COLUMN_DEFS = [
    # (header, key, width, group)
    ("case_number", "case_number", 18, "data"),
    ("date", "date", 12, "data"),
    ("court", "court", 30, "data"),
    ("originating_court", "originating_court", 25, "data"),
    ("is_appeal", "is_appeal", 10, "data"),
    ("risk_label", "risk_label", 12, "data"),
    ("application_outcome", "application_outcome", 20, "data"),
    ("application_outcome_sv", "application_outcome_sv", 25, "data"),
    ("power_plant_name", "power_plant_name", 25, "data"),
    ("watercourse", "watercourse", 20, "data"),
    ("operator_name", "operator_name", 25, "data"),
    ("total_cost_sek", "total_cost_sek", 15, "data"),
    ("cost_source_text", "cost_source_text", 40, "data"),
    ("cost_original_text", "cost_original_text", 20, "data"),
    ("processing_time_days", "processing_time_days", 18, "data"),
    ("measures_ordered", "measures_ordered_str", 30, "data"),
    ("domslut_measures", "domslut_measures_str", 30, "data"),
    ("extracted_measures", "extracted_measures_str", 30, "data"),
    # Strommen cross-reference
    ("STR_matched", "STR_matched", 12, "str"),
    ("STR_plant_name", "STR_plant_name", 25, "str"),
    ("STR_court", "STR_court", 15, "str"),
    ("STR_beslut", "STR_beslut", 20, "str"),
    ("STR_size", "STR_size", 22, "str"),
    ("STR_provningsgrupp", "STR_provningsgrupp", 20, "str"),
    ("STR_lan", "STR_lan", 18, "str"),
    ("STR_huvudavrinningsomrade", "STR_huvudavrinningsomrade", 22, "str"),
    ("STR_status", "STR_status", 20, "str"),
    ("STR_faunapassage", "STR_faunapassage", 18, "str"),
    ("STR_fingaller", "STR_fingaller", 18, "str"),
    ("STR_minimitappning", "STR_minimitappning", 22, "str"),
    # Validation flags
    ("FLAG_cost_over_5M", "FLAG_cost_over_5M", 15, "flag"),
    ("FLAG_outcome_unclear", "FLAG_outcome_unclear", 18, "flag"),
    ("FLAG_no_measures", "FLAG_no_measures", 15, "flag"),
    ("FLAG_no_cost", "FLAG_no_cost", 12, "flag"),
    ("FLAG_no_plant_name", "FLAG_no_plant_name", 17, "flag"),
    ("FLAG_no_operator", "FLAG_no_operator", 15, "flag"),
    ("FLAG_no_processing_time", "FLAG_no_processing_time", 20, "flag"),
    ("FLAG_unlabeled", "FLAG_unlabeled", 14, "flag"),
    ("FLAG_plant_name_prefix", "FLAG_plant_name_prefix", 20, "flag"),
    ("FLAG_outcome_mismatch", "FLAG_outcome_mismatch", 20, "flag"),
    # Review columns (empty for human reviewer)
    ("REVIEW_outcome_correct", "REVIEW_outcome_correct", 22, "review"),
    ("REVIEW_corrected_outcome", "REVIEW_corrected_outcome", 22, "review"),
    ("REVIEW_cost_correct", "REVIEW_cost_correct", 18, "review"),
    ("REVIEW_corrected_cost", "REVIEW_corrected_cost", 18, "review"),
    ("REVIEW_measures_correct", "REVIEW_measures_correct", 20, "review"),
    ("REVIEW_risk_label_correct", "REVIEW_risk_label_correct", 22, "review"),
    ("REVIEW_notes", "REVIEW_notes", 30, "review"),
    # Source text excerpts
    ("domslut_excerpt", "domslut_excerpt", 50, "excerpt"),
    ("bakgrund_excerpt", "bakgrund_excerpt", 50, "excerpt"),
]

FLAG_KEYS = [c[1] for c in COLUMN_DEFS if c[3] == "flag"]
REVIEW_KEYS = [c[1] for c in COLUMN_DEFS if c[3] == "review"]


def build_row_data(decision, label_lookup, strommen_lookup):
    """Build a flat dict of all column values for one decision."""
    meta = decision.get("metadata", {})
    dec_id = decision.get("id", "")

    # Label info
    label_info = label_lookup.get(dec_id, {})
    risk_label = label_info.get("label", "")
    scoring = label_info.get("scoring_details", {})

    # Measures
    measures_ordered_list = meta.get("measures_ordered", []) or []
    measures_ordered_str = ", ".join(
        m.get("type", "") for m in measures_ordered_list if m.get("type")
    )

    domslut_measures = scoring.get("domslut_measures", []) or []
    domslut_measures_str = ", ".join(domslut_measures)

    extracted_measures = decision.get("extracted_measures", []) or []
    extracted_measures_str = ", ".join(extracted_measures)

    # Costs
    extracted_costs = decision.get("extracted_costs", []) or []
    cost_source_text = ""
    cost_original_text = ""
    if extracted_costs:
        ctx = extracted_costs[0].get("context", "")
        cost_source_text = ctx[:120] if ctx else ""
        cost_original_text = extracted_costs[0].get("original", "")

    # Sections
    sections = decision.get("sections", {}) or {}
    domslut_text = sections.get("domslut", "") or ""
    bakgrund_text = sections.get("bakgrund", "") or ""

    # Strommen cross-reference
    case_number = meta.get("case_number", "")
    ncn = normalize_case_number(case_number)
    str_match = strommen_lookup.get(ncn)

    row = {
        "id": dec_id,
        "case_number": case_number,
        "date": meta.get("date", ""),
        "court": meta.get("court", ""),
        "originating_court": meta.get("originating_court", ""),
        "is_appeal": meta.get("is_appeal", False),
        "risk_label": risk_label,
        "application_outcome": meta.get("application_outcome", ""),
        "application_outcome_sv": meta.get("application_outcome_sv", ""),
        "power_plant_name": meta.get("power_plant_name"),
        "watercourse": meta.get("watercourse", ""),
        "operator_name": meta.get("operator_name"),
        "total_cost_sek": meta.get("total_cost_sek"),
        "cost_source_text": cost_source_text,
        "cost_original_text": cost_original_text,
        "processing_time_days": meta.get("processing_time_days"),
        "measures_ordered_str": measures_ordered_str,
        "domslut_measures_str": domslut_measures_str,
        "extracted_measures_str": extracted_measures_str,
        # Strommen
        "STR_matched": str_match is not None,
        "STR_plant_name": "",
        "STR_court": "",
        "STR_beslut": "",
        "STR_size": "",
        "STR_provningsgrupp": "",
        "STR_lan": "",
        "STR_huvudavrinningsomrade": "",
        "STR_status": "",
        "STR_faunapassage": "",
        "STR_fingaller": "",
        "STR_minimitappning": "",
        # Excerpts
        "domslut_excerpt": domslut_text[:2000],
        "bakgrund_excerpt": bakgrund_text[:1000],
        # Review (empty)
        "REVIEW_outcome_correct": "",
        "REVIEW_corrected_outcome": "",
        "REVIEW_cost_correct": "",
        "REVIEW_corrected_cost": "",
        "REVIEW_measures_correct": "",
        "REVIEW_risk_label_correct": "",
        "REVIEW_notes": "",
    }

    if str_match:
        sr = str_match["row"]
        row["STR_plant_name"] = sr.get("Vattenkraftverk", "")
        row["STR_court"] = sr.get("Domstol", "")
        row["STR_beslut"] = sr.get("Beslut", "")
        row["STR_size"] = (sr.get("Vattenkraftverk storlek", "") or "").strip()
        row["STR_provningsgrupp"] = sr.get("Prövningsgrupp", "")
        row["STR_lan"] = sr.get("Län", "")
        row["STR_huvudavrinningsomrade"] = sr.get(
            "Huvudavrinningsområde", ""
        )
        row["STR_status"] = sr.get("Status samverkansprocess", "")
        row["STR_faunapassage"] = sr.get("Faunapassage uppströms", "")
        row["STR_fingaller"] = sr.get("Fingaller framför turbinintag", "")
        row["STR_minimitappning"] = sr.get(
            "Tappningsbestämmelse minimitappningar", ""
        )

    # Flags
    flags = compute_flags(row, str_match, label_info)
    row.update(flags)

    return row


def write_datavalidering_sheet(ws, rows):
    """Write Sheet 1: Datavalidering."""
    # Fills
    red_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC",
                           fill_type="solid")
    green_fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC",
                             fill_type="solid")
    yellow_fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC",
                              fill_type="solid")
    blue_fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF",
                            fill_type="solid")
    bold_font = Font(bold=True)

    # Headers
    for col_idx, (header, key, width, group) in enumerate(COLUMN_DEFS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = bold_font
        ws.column_dimensions[get_column_letter(col_idx)].width = min(width, 60)

        if group == "str":
            cell.fill = blue_fill
        elif group == "review":
            cell.fill = yellow_fill

    # Data rows
    for row_idx, row_data in enumerate(rows, 2):
        for col_idx, (header, key, width, group) in enumerate(COLUMN_DEFS, 1):
            value = row_data.get(key, "")
            if value is None:
                value = ""

            cell = ws.cell(row=row_idx, column=col_idx, value=value)

            # Flag columns: red fill if TRUE
            if group == "flag" and value is True:
                cell.fill = red_fill

            # Review columns: yellow fill
            if group == "review":
                cell.fill = yellow_fill

            # risk_label coloring
            if key == "risk_label":
                if value == "HIGH_RISK":
                    cell.fill = red_fill
                elif value == "LOW_RISK":
                    cell.fill = green_fill

            # Wrap text on excerpts
            if group == "excerpt":
                cell.alignment = Alignment(wrap_text=True)

    # Freeze pane + auto-filter
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMN_DEFS))}1"


def write_sammanfattning_sheet(ws, rows, strommen_lookup):
    """Write Sheet 2: Sammanfattning."""
    bold_font = Font(bold=True)

    total = len(rows)
    labeled = [r for r in rows if r.get("risk_label")]
    unlabeled = [r for r in rows if not r.get("risk_label")]
    unclear = [r for r in rows if r.get("application_outcome") == "unclear"]
    matched = [r for r in rows if r.get("STR_matched")]

    # Outcome mismatches with details
    mismatches = []
    for r in rows:
        if r.get("FLAG_outcome_mismatch"):
            mismatches.append(
                f"{r['case_number']} ({r['application_outcome']} -> "
                f"{r['STR_beslut']})"
            )

    no_cost = [r for r in rows if r.get("FLAG_no_cost")]
    high_cost = [r for r in rows if r.get("FLAG_cost_over_5M")]
    no_measures = [r for r in rows if r.get("FLAG_no_measures")]
    no_plant = [r for r in rows if r.get("FLAG_no_plant_name")]
    no_operator = [r for r in rows if r.get("FLAG_no_operator")]
    no_time = [r for r in rows if r.get("FLAG_no_processing_time")]
    bad_prefix = [r for r in rows if r.get("FLAG_plant_name_prefix")]

    appeals = [r for r in rows if r.get("is_appeal")]
    first_instance = [r for r in rows if not r.get("is_appeal")]

    data_rows = [
        ("Totalt antal beslut:", total),
        ("Märkta (labeled):", len(labeled)),
        (
            "Omärkta:",
            f"{len(unlabeled)} ({', '.join(r['case_number'] for r in unlabeled)})"
            if unlabeled else "0",
        ),
        (
            'Utfall "unclear":',
            f"{len(unclear)} ({', '.join(r['case_number'] for r in unclear)})"
            if unclear else "0",
        ),
        ("Matchade i Strömmen:", len(matched)),
        (
            "Utfallsmismatch vs Strömmen:",
            f"{len(mismatches)} ({'; '.join(mismatches)})"
            if mismatches else "0",
        ),
        ("Saknar kostnad:", len(no_cost)),
        (
            "Misstänkt kostnad (>5M):",
            "{} ({})".format(
                len(high_cost),
                "; ".join(
                    "{}: {}".format(r["case_number"], r["total_cost_sek"])
                    for r in high_cost
                ),
            )
            if high_cost else "0",
        ),
        ("Saknar åtgärder (alla fält):", len(no_measures)),
        ("Saknar kraftverksnamn:", len(no_plant)),
        ("Saknar operatör:", len(no_operator)),
        ("Saknar handläggningstid:", len(no_time)),
        (
            "Felaktigt kraftverksnamn (prefix):",
            f"{len(bad_prefix)} ({', '.join(r['case_number'] for r in bad_prefix)})"
            if bad_prefix else "0",
        ),
        ("", ""),
        ("MÖD-överklaganden i vår databas:", len(appeals)),
        ("Första instans i vår databas:", len(first_instance)),
        ("", ""),
        ("Strömmen-källa:", "uppdaterad 2025-07-02"),
        ("Rapport genererad:", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
    ]

    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 80

    for row_idx, (label, value) in enumerate(data_rows, 1):
        cell_a = ws.cell(row=row_idx, column=1, value=label)
        if label:
            cell_a.font = bold_font
        ws.cell(row=row_idx, column=2, value=value)


def write_nap_landskapet_sheet(ws, strommen_file1, strommen_file2, rows):
    """Write Sheet 3: NAP_landskapet."""
    bold_font = Font(bold=True)
    all_str = strommen_file1 + strommen_file2
    total_nap = len(all_str)

    # Status counts from file 1
    pagar = sum(
        1 for r in strommen_file1
        if (r.get("Status samverkansprocess") or "").strip() == "Pågår"
    )
    avslutad = sum(
        1 for r in strommen_file1
        if (r.get("Status samverkansprocess") or "").strip() == "Avslutad"
    )
    ej_pabor = sum(
        1 for r in strommen_file2
        if (r.get("Status samverkansprocess") or "").strip() == "Ej påbörjad"
    )

    # Per domstol
    court_counts = {}
    for r in all_str:
        c = (r.get("Domstol") or "").strip()
        if c:
            court_counts[c] = court_counts.get(c, 0) + 1
    court_sorted = sorted(court_counts.items(), key=lambda x: -x[1])

    # Per storlek
    size_counts = {}
    for r in all_str:
        s = (r.get("Vattenkraftverk storlek") or "").strip()
        if s:
            size_counts[s] = size_counts.get(s, 0) + 1

    # Avgjorda arenden (File 1, Beslut != "Ej Beslut")
    avgjorda = [
        r for r in strommen_file1
        if (r.get("Beslut") or "").strip()
        and (r.get("Beslut") or "").strip() != "Ej Beslut"
    ]
    # Unique case numbers among avgjorda
    avgjorda_cn = set()
    for r in avgjorda:
        cn = (r.get("Målnummer") or "").strip()
        if cn:
            avgjorda_cn.add(cn)
    avgjorda_unique = len(avgjorda_cn)

    beslut_types = {}
    for r in avgjorda:
        b = (r.get("Beslut") or "").strip()
        beslut_types[b] = beslut_types.get(b, 0) + 1

    # Our coverage
    matched_count = sum(1 for r in rows if r.get("STR_matched"))
    mod_miljovillkor_avgjorda = sum(
        1 for r in avgjorda
        if (r.get("Beslut") or "").strip() == "Moderna miljövillkor"
    )

    # Avgjorda we DON'T have — build set of our normalized case numbers
    our_ncns = set()
    for r in rows:
        ncn = normalize_case_number(r.get("case_number", ""))
        if ncn:
            our_ncns.add(ncn)

    avgjorda_str_cns = set()
    for r in avgjorda:
        cn = (r.get("Målnummer") or "").strip()
        if cn:
            avgjorda_str_cns.add(normalize_case_number(cn))

    missing = avgjorda_str_cns - our_ncns
    missing_rows = [
        r for r in avgjorda
        if normalize_case_number((r.get("Målnummer") or "").strip()) in missing
    ]
    # De-duplicate missing rows by normalized case number
    seen_missing = set()
    missing_unique = []
    for r in missing_rows:
        ncn = normalize_case_number((r.get("Målnummer") or "").strip())
        if ncn not in seen_missing:
            seen_missing.add(ncn)
            missing_unique.append(r)

    missing_by_type = {}
    for r in missing_unique:
        b = (r.get("Beslut") or "").strip()
        missing_by_type[b] = missing_by_type.get(b, 0) + 1

    mod_appeals = sum(1 for r in rows if r.get("is_appeal") and not r.get("STR_matched"))

    # Write content
    content_lines = [
        ("=== TOTAL NAP SCOPE ===", True),
        (f"Total kraftverk i NAP: {total_nap}", False),
        (f"Pågår (File 1): {pagar}", False),
        (f"Avslutad (File 1): {avslutad}", False),
        (f"Ej påbörjad (File 2): {ej_pabor}", False),
        ("", False),
        ("=== PER DOMSTOL (both files combined) ===", True),
    ]
    for court, cnt in court_sorted:
        content_lines.append((f"  {court}: {cnt}", False))

    content_lines.append(("", False))
    content_lines.append(("=== PER STORLEK (both files combined) ===", True))
    for sz, cnt in sorted(size_counts.items()):
        content_lines.append((f"  {sz}: {cnt}", False))

    content_lines.append(("", False))
    content_lines.append(
        ('=== AVGJORDA ÄRENDEN (File 1, Beslut != "Ej Beslut") ===', True)
    )
    content_lines.append(
        (f"Total avgjorda: {avgjorda_unique} (unika målnummer)", False)
    )
    content_lines.append(("Per typ:", True))
    for btype in [
        "Moderna miljövillkor",
        "Återkallande (utrivning)",
        "Avvisning",
        "Avslag",
        "Avskrivet",
    ]:
        cnt = beslut_types.get(btype, 0)
        content_lines.append((f"  {btype}: {cnt}", False))
    # Any other types not in the standard list
    for btype, cnt in sorted(beslut_types.items()):
        if btype not in [
            "Moderna miljövillkor",
            "Återkallande (utrivning)",
            "Avvisning",
            "Avslag",
            "Avskrivet",
        ]:
            content_lines.append((f"  {btype}: {cnt}", False))

    content_lines.append(("", False))
    content_lines.append(("=== VÅR TÄCKNING ===", True))
    content_lines.append(
        (f"Våra beslut i Strömmen: {matched_count}", False)
    )
    content_lines.append(
        (f"Avgjorda i Strömmen vi SAKNAR: {len(missing_unique)}", False)
    )
    content_lines.append(
        (f"  varav Moderna miljövillkor: {missing_by_type.get('Moderna miljövillkor', 0)}"
         " (potentiellt mest värdefulla att inhämta)", False)
    )
    content_lines.append(
        (f"  varav Återkallande: {missing_by_type.get('Återkallande (utrivning)', 0)}",
         False)
    )
    content_lines.append(
        (f"  varav Avvisning: {missing_by_type.get('Avvisning', 0)}", False)
    )
    avslag_avskrivet = (
        missing_by_type.get("Avslag", 0)
        + missing_by_type.get("Avskrivet", 0)
    )
    content_lines.append(
        (f"  varav Avslag/Avskrivet: {avslag_avskrivet}", False)
    )

    content_lines.append(("", False))
    content_lines.append(
        (f"Våra MÖD-beslut (ej i Strömmen): {mod_appeals}", False)
    )
    content_lines.append(
        ("(MÖD-överklaganden har egna målnummer och listas "
         "ej i Strömmen)", False)
    )

    ws.column_dimensions["A"].width = 80

    for row_idx, (text, is_bold) in enumerate(content_lines, 1):
        cell = ws.cell(row=row_idx, column=1, value=text)
        if is_bold:
            cell.font = bold_font


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(
        description="Generate data validation report for NAP court decisions"
    )
    parser.add_argument(
        "--data-root",
        default=None,
        help="Path to the Data directory (default: auto-detect)",
    )
    args = parser.parse_args()

    data_root = resolve_data_root(args.data_root)
    print(f"Data root: {data_root}")

    # Load cleaned court texts
    ct_path = os.path.join(data_root, "processed", "cleaned_court_texts.json")
    print(f"Loading {ct_path} ...")
    ct_data = load_json(ct_path)
    decisions = ct_data.get("decisions", [])
    print(f"  Loaded {len(decisions)} decisions")

    # Load labeled dataset
    lb_path = os.path.join(data_root, "processed", "labeled_dataset_binary.json")
    print(f"Loading {lb_path} ...")
    lb_data = load_json(lb_path)
    label_lookup = build_label_lookup(lb_data)
    print(f"  Built label lookup for {len(label_lookup)} decisions")

    # Load Strommen files
    str_dir = os.path.join(data_root, "Strommen")
    # Find the two CSV files — handle trailing spaces in filenames
    str_files = []
    for f in sorted(os.listdir(str_dir)):
        if f.lower().endswith(".csv"):
            str_files.append(os.path.join(str_dir, f))

    if len(str_files) < 2:
        print(f"ERROR: Expected 2 CSV files in {str_dir}, found {len(str_files)}")
        sys.exit(1)

    # File without "(1)" is File 1 (ongoing/completed), file with "(1)" is File 2
    file1_path = None
    file2_path = None
    for fp in str_files:
        basename = os.path.basename(fp)
        if "(1)" in basename:
            file2_path = fp
        else:
            file1_path = fp

    if not file1_path or not file2_path:
        # Fallback: first file is file1, second is file2
        file1_path = str_files[0]
        file2_path = str_files[1]

    print(f"Loading Strömmen File 1: {os.path.basename(file1_path)} ...")
    strommen_file1 = load_strommen_csv(file1_path)
    print(f"  Loaded {len(strommen_file1)} rows")

    print(f"Loading Strömmen File 2: {os.path.basename(file2_path)} ...")
    strommen_file2 = load_strommen_csv(file2_path)
    print(f"  Loaded {len(strommen_file2)} rows")

    # Build Strommen lookup from both files
    all_strommen = strommen_file1 + strommen_file2
    strommen_lookup = build_strommen_lookup(all_strommen)
    print(f"  Built Strömmen lookup with {len(strommen_lookup)} unique case numbers")

    # Build rows
    print("\nBuilding validation data ...")
    all_rows = []
    for dec in decisions:
        row = build_row_data(dec, label_lookup, strommen_lookup)
        all_rows.append(row)

    # Sort by case_number
    all_rows.sort(key=lambda r: r.get("case_number", ""))

    # Create workbook
    wb = Workbook()

    # Sheet 1: Datavalidering
    ws1 = wb.active
    ws1.title = "Datavalidering"
    write_datavalidering_sheet(ws1, all_rows)

    # Sheet 2: Sammanfattning
    ws2 = wb.create_sheet("Sammanfattning")
    write_sammanfattning_sheet(ws2, all_rows, strommen_lookup)

    # Sheet 3: NAP_landskapet
    ws3 = wb.create_sheet("NAP_landskapet")
    write_nap_landskapet_sheet(ws3, strommen_file1, strommen_file2, all_rows)

    # Output
    out_dir = os.path.join(data_root, "validation")
    os.makedirs(out_dir, exist_ok=True)
    out_path = os.path.join(out_dir, "data_validation_report.xlsx")
    wb.save(out_path)

    # Summary
    flagged = sum(
        1
        for r in all_rows
        if any(r.get(fk) for fk in FLAG_KEYS)
    )
    matched = sum(1 for r in all_rows if r.get("STR_matched"))
    labeled = sum(1 for r in all_rows if r.get("risk_label"))

    print(f"\n{'='*60}")
    print(f"REPORT GENERATED: {out_path}")
    print(f"{'='*60}")
    print(f"  Total decisions:       {len(all_rows)}")
    print(f"  Labeled:               {labeled}")
    print(f"  Unlabeled:             {len(all_rows) - labeled}")
    print(f"  Matched in Strömmen:   {matched}")
    print(f"  Flagged for review:    {flagged}")
    print(f"  Appeals (MÖD):         {sum(1 for r in all_rows if r.get('is_appeal'))}")
    print(f"  First instance:        {sum(1 for r in all_rows if not r.get('is_appeal'))}")
    print(f"  NAP total (Strömmen):  {len(all_strommen)}")
    print(f"{'='*60}")


if __name__ == "__main__":
    main()
