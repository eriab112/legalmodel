"""
Generate a data validation report as an Excel spreadsheet.

Reads cleaned_court_texts.json and labeled_dataset_binary.json,
produces Data/validation/data_validation_report.xlsx with:
  - Sheet "Datavalidering": one row per decision with extracted metadata,
    source text excerpts, validation flags, and human-review columns.
  - Sheet "Sammanfattning": summary statistics.
"""

import json
import os
import sys

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule

# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------
BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
CLEANED_PATH = os.path.join(BASE_DIR, "Data", "processed", "cleaned_court_texts.json")
LABELED_PATH = os.path.join(BASE_DIR, "Data", "processed", "labeled_dataset_binary.json")
OUTPUT_DIR = os.path.join(BASE_DIR, "Data", "validation")
OUTPUT_PATH = os.path.join(OUTPUT_DIR, "data_validation_report.xlsx")


def load_json(path: str) -> dict:
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)


def build_label_map(labeled: dict) -> dict[str, str]:
    """Map decision id -> risk_label from the labeled dataset."""
    label_map: dict[str, str] = {}
    for split in labeled.get("splits", {}).values():
        for entry in split:
            label_map[entry["id"]] = entry.get("label", "")
    return label_map


def build_scoring_map(labeled: dict) -> dict[str, dict]:
    """Map decision id -> scoring_details from the labeled dataset."""
    scoring_map: dict[str, dict] = {}
    for split in labeled.get("splits", {}).values():
        for entry in split:
            if "scoring_details" in entry:
                scoring_map[entry["id"]] = entry["scoring_details"]
    return scoring_map


def measures_ordered_str(measures) -> str:
    """Convert measures_ordered list to a JSON string."""
    if not measures:
        return ""
    return json.dumps(measures, ensure_ascii=False)


def measures_list_str(measures) -> str:
    """Convert a simple list of strings to comma-separated."""
    if not measures:
        return ""
    return ", ".join(str(m) for m in measures)


def cost_source_text(decision: dict) -> str:
    """Return the context field from the first extracted_costs entry."""
    costs = decision.get("extracted_costs", [])
    if costs and isinstance(costs, list) and len(costs) > 0:
        return costs[0].get("context", "")
    return ""


def truncate(text: str | None, max_len: int) -> str:
    if not text:
        return ""
    if len(text) <= max_len:
        return text
    return text[:max_len] + "..."


def main():
    # Load data
    cleaned = load_json(CLEANED_PATH)
    labeled = load_json(LABELED_PATH)

    decisions = cleaned.get("decisions", [])
    label_map = build_label_map(labeled)
    scoring_map = build_scoring_map(labeled)

    # Sort by case_number
    decisions.sort(key=lambda d: d.get("metadata", {}).get("case_number", ""))

    # Ensure output directory exists
    os.makedirs(OUTPUT_DIR, exist_ok=True)

    # Create workbook
    wb = Workbook()

    # -----------------------------------------------------------------------
    # Sheet 1: Datavalidering
    # -----------------------------------------------------------------------
    ws = wb.active
    ws.title = "Datavalidering"

    # Define columns
    headers = [
        # Extracted data
        "case_number",
        "date",
        "court",
        "originating_court",
        "is_appeal",
        "risk_label",
        "application_outcome",
        "application_outcome_sv",
        "power_plant_name",
        "watercourse",
        "operator_name",
        "total_cost_sek",
        "cost_source_text",
        "processing_time_days",
        "measures_ordered",
        "domslut_measures",
        "extracted_measures",
        # Source text
        "domslut_excerpt",
        "bakgrund_excerpt",
        # Validation flags
        "FLAG_cost_suspicious",
        "FLAG_outcome_unclear",
        "FLAG_no_measures",
        "FLAG_no_cost",
        "FLAG_no_plant_name",
        "FLAG_no_operator",
        "FLAG_no_processing_time",
        "FLAG_unlabeled",
        # Review columns
        "REVIEW_outcome_correct",
        "REVIEW_cost_correct",
        "REVIEW_measures_correct",
        "REVIEW_risk_label_correct",
        "REVIEW_notes",
    ]

    # Text columns that should be capped at 50 chars width
    text_columns = {
        "cost_source_text", "measures_ordered", "domslut_measures",
        "extracted_measures", "domslut_excerpt", "bakgrund_excerpt",
        "REVIEW_notes",
    }

    # Write header row
    bold_font = Font(bold=True)
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = bold_font

    # Freeze the header row
    ws.freeze_panes = "A2"

    # Write data rows
    for row_idx, decision in enumerate(decisions, start=2):
        meta = decision.get("metadata", {})
        sections = decision.get("sections", {})
        dec_id = decision.get("id", "")

        risk_label = label_map.get(dec_id, "")
        scoring = scoring_map.get(dec_id, {})

        # Extracted data columns
        case_number = meta.get("case_number", "")
        date = meta.get("date", "")
        court = meta.get("court", "")
        originating_court = meta.get("originating_court", "")
        is_appeal = meta.get("is_appeal", "")
        application_outcome = meta.get("application_outcome", "")
        application_outcome_sv = meta.get("application_outcome_sv", "")
        power_plant_name = meta.get("power_plant_name")
        watercourse = meta.get("watercourse", "")
        operator_name = meta.get("operator_name")
        total_cost_sek = meta.get("total_cost_sek")
        cost_src = cost_source_text(decision)
        processing_time_days = meta.get("processing_time_days")
        measures_ord = meta.get("measures_ordered", [])
        domslut_meas = scoring.get("domslut_measures", [])
        extracted_meas = decision.get("extracted_measures", [])

        # Source text excerpts
        domslut_excerpt = truncate(sections.get("domslut", ""), 1500)
        bakgrund_excerpt = truncate(sections.get("bakgrund", ""), 800)

        # Validation flags
        flag_cost_suspicious = (
            total_cost_sek is not None and total_cost_sek > 5_000_000
        )
        flag_outcome_unclear = application_outcome == "unclear"
        flag_no_measures = (
            not measures_ord and not domslut_meas and not extracted_meas
        )
        flag_no_cost = total_cost_sek is None
        flag_no_plant_name = power_plant_name is None
        flag_no_operator = operator_name is None
        flag_no_processing_time = processing_time_days is None
        flag_unlabeled = risk_label == ""

        row_data = [
            case_number,
            date,
            court,
            originating_court,
            is_appeal,
            risk_label,
            application_outcome,
            application_outcome_sv,
            power_plant_name if power_plant_name is not None else "",
            watercourse,
            operator_name if operator_name is not None else "",
            total_cost_sek if total_cost_sek is not None else "",
            cost_src,
            processing_time_days if processing_time_days is not None else "",
            measures_ordered_str(measures_ord),
            measures_list_str(domslut_meas),
            measures_list_str(extracted_meas),
            domslut_excerpt,
            bakgrund_excerpt,
            flag_cost_suspicious,
            flag_outcome_unclear,
            flag_no_measures,
            flag_no_cost,
            flag_no_plant_name,
            flag_no_operator,
            flag_no_processing_time,
            flag_unlabeled,
            "",  # REVIEW_outcome_correct
            "",  # REVIEW_cost_correct
            "",  # REVIEW_measures_correct
            "",  # REVIEW_risk_label_correct
            "",  # REVIEW_notes
        ]

        for col_idx, value in enumerate(row_data, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    # -----------------------------------------------------------------------
    # Auto-fit column widths (capped at 50 for text columns)
    # -----------------------------------------------------------------------
    for col_idx, header in enumerate(headers, start=1):
        col_letter = get_column_letter(col_idx)
        max_width = len(header)
        for row in range(2, ws.max_row + 1):
            cell_value = ws.cell(row=row, column=col_idx).value
            if cell_value is not None:
                max_width = max(max_width, len(str(cell_value)))

        cap = 50 if header in text_columns else 80
        col_width = min(max_width + 2, cap)
        ws.column_dimensions[col_letter].width = col_width

    # -----------------------------------------------------------------------
    # Conditional formatting
    # -----------------------------------------------------------------------
    last_row = ws.max_row
    red_fill = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")
    green_fill = PatternFill(start_color="99FF99", end_color="99FF99", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")

    # FLAG columns: red fill if TRUE
    flag_cols = [
        "FLAG_cost_suspicious", "FLAG_outcome_unclear", "FLAG_no_measures",
        "FLAG_no_cost", "FLAG_no_plant_name", "FLAG_no_operator",
        "FLAG_no_processing_time", "FLAG_unlabeled",
    ]
    for flag_col in flag_cols:
        col_idx = headers.index(flag_col) + 1
        col_letter = get_column_letter(col_idx)
        cell_range = f"{col_letter}2:{col_letter}{last_row}"
        ws.conditional_formatting.add(
            cell_range,
            CellIsRule(operator="equal", formula=["TRUE"], fill=red_fill),
        )

    # Risk label: red for HIGH_RISK, green for LOW_RISK
    risk_col_idx = headers.index("risk_label") + 1
    risk_col_letter = get_column_letter(risk_col_idx)
    risk_range = f"{risk_col_letter}2:{risk_col_letter}{last_row}"
    ws.conditional_formatting.add(
        risk_range,
        CellIsRule(operator="equal", formula=['"HIGH_RISK"'], fill=red_fill),
    )
    ws.conditional_formatting.add(
        risk_range,
        CellIsRule(operator="equal", formula=['"LOW_RISK"'], fill=green_fill),
    )

    # REVIEW columns: yellow fill (always)
    review_cols = [
        "REVIEW_outcome_correct", "REVIEW_cost_correct",
        "REVIEW_measures_correct", "REVIEW_risk_label_correct",
        "REVIEW_notes",
    ]
    for review_col in review_cols:
        col_idx = headers.index(review_col) + 1
        col_letter = get_column_letter(col_idx)
        for row in range(2, last_row + 1):
            ws.cell(row=row, column=col_idx).fill = yellow_fill

    # Wrap text for excerpt columns
    wrap_alignment = Alignment(wrap_text=True, vertical="top")
    for excerpt_col in ["domslut_excerpt", "bakgrund_excerpt"]:
        col_idx = headers.index(excerpt_col) + 1
        for row in range(1, last_row + 1):
            ws.cell(row=row, column=col_idx).alignment = wrap_alignment

    # -----------------------------------------------------------------------
    # Sheet 2: Sammanfattning (Summary)
    # -----------------------------------------------------------------------
    ws2 = wb.create_sheet("Sammanfattning")
    ws2.column_dimensions["A"].width = 35
    ws2.column_dimensions["B"].width = 15

    # Compute summary stats
    total_decisions = len(decisions)
    labeled_count = sum(1 for d in decisions if label_map.get(d.get("id", ""), "") != "")
    unlabeled_count = total_decisions - labeled_count
    unclear_outcome = sum(
        1 for d in decisions
        if d.get("metadata", {}).get("application_outcome") == "unclear"
    )
    missing_cost = sum(
        1 for d in decisions
        if d.get("metadata", {}).get("total_cost_sek") is None
    )
    suspicious_cost = sum(
        1 for d in decisions
        if (d.get("metadata", {}).get("total_cost_sek") or 0) > 5_000_000
    )
    missing_measures = sum(
        1 for d in decisions
        if (
            not d.get("metadata", {}).get("measures_ordered")
            and not scoring_map.get(d.get("id", ""), {}).get("domslut_measures")
            and not d.get("extracted_measures")
        )
    )
    missing_plant_name = sum(
        1 for d in decisions
        if d.get("metadata", {}).get("power_plant_name") is None
    )
    missing_operator = sum(
        1 for d in decisions
        if d.get("metadata", {}).get("operator_name") is None
    )
    missing_processing_time = sum(
        1 for d in decisions
        if d.get("metadata", {}).get("processing_time_days") is None
    )

    summary_rows = [
        ("Statistic", "Value"),
        ("Total decisions", total_decisions),
        ("Labeled", labeled_count),
        ("Unlabeled", unlabeled_count),
        ("Unclear outcome", unclear_outcome),
        ("Missing cost", missing_cost),
        ("Suspicious cost (>5M)", suspicious_cost),
        ("Missing measures", missing_measures),
        ("Missing plant name", missing_plant_name),
        ("Missing operator", missing_operator),
        ("Missing processing time", missing_processing_time),
    ]

    for row_idx, (label, value) in enumerate(summary_rows, start=1):
        cell_a = ws2.cell(row=row_idx, column=1, value=label)
        cell_b = ws2.cell(row=row_idx, column=2, value=value)
        if row_idx == 1:
            cell_a.font = bold_font
            cell_b.font = bold_font

    ws2.freeze_panes = "A2"

    # -----------------------------------------------------------------------
    # Save
    # -----------------------------------------------------------------------
    wb.save(OUTPUT_PATH)
    print(f"Validation report saved to: {OUTPUT_PATH}")

    # Print summary
    print(f"\n--- Summary ---")
    for label, value in summary_rows[1:]:
        print(f"  {label}: {value}")


if __name__ == "__main__":
    main()
